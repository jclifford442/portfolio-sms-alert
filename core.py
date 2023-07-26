# -*- coding: utf-8 -*-
import datetime
from openpyxl import load_workbook, Workbook
from twilio.rest import Client
import helpers
import config



def first_run():
    if not helpers.check_file("../sms-stock-alert/files/portfolio_holdings.xlsx"):
        helpers.build_holdings_worksheet()
    if not helpers.check_file("../sms-stock-alert/files/portfolio_history.xlsx"):
        helpers.build_history_worksheet()


def daily_gain_amount():
    file_path = "../sms-stock-alert/files/portfolio_history.xlsx"
    wb = load_workbook(file_path)
    ws = wb.active
    yesterday_total = ws.cell(row=ws.max_row, column=2).value
    today_total = helpers.get_current_day_portfolio_value()
    ws.append([datetime.date.today(), today_total])
    wb.save("../sms-stock-alert/files/portfolio_history.xlsx")
    wb.close()
    if today_total - yesterday_total > 0.1 or today_total - yesterday_total < -0.1:
        send_text(today_total - yesterday_total)


def send_text(amount):
    client = Client(config.account_sid, config.auth_token)
    total = round(amount, 2)
    message = client.messages.create(
        from_=config.from_num,
        body='Daily gain/loss: ' + "${:,.2f}".format(total),
        to=config.to_num
        )