import yfinance as yf
from openpyxl import load_workbook, Workbook
import datetime


def get_price(share_count, ticker):
    total = share_count * yf.Ticker(ticker).fast_info['lastPrice']
    return total


def get_current_day_portfolio_value():
    wb = load_workbook("../sms-stock-alert/files/portfolio_holdings.xlsx")
    ws = wb.active
    daily_total = 0
    for row in ws.iter_rows():
        ticker = row[0].value
        share_count = float(row[1].value)
        daily_total += get_price(share_count, ticker)
    wb.close()
    return daily_total


#Helper Function
def check_valid_ticker(asset: str) -> bool:
    """
    Checks if an asset is available via the Yahoo Finance API.
    """
    info = yf.Ticker(asset).history(
        period='7d',
        interval='1d')
    return len(info) > 0


def check_file(file_path: str) -> bool:
    """
    Checks if a file exists.
    """
    try:
        open(file_path)
        return True
    except FileNotFoundError:
        return False



def build_holdings_worksheet():
    wb = Workbook()
    ws = wb.active
    print("Enter the ticker symbol (ie. aapl) and number of shares you want to add to your portfolio tracker\n"
          "When finished, press 'c' followed by enter")
    while True:
        ticker = input(f'Enter ticker symbol: ').lower()
        if ticker == "c":
            break
        elif not check_valid_ticker(ticker):
            ticker = input(f'Enter ticker symbol: ').lower()
        else:
            quantity = input(f'Enter number of shares: ')
            data = [ticker, quantity]
            ws.append(data)
    wb.save("../sms-stock-alert/files/portfolio_holdings.xlsx")
    wb.close()


def build_history_worksheet():
    wb = Workbook()
    ws = wb.active
    ws.append([datetime.date.today(), get_current_day_portfolio_value()])
    wb.save("../sms-stock-alert/files/portfolio_history.xlsx")
    wb.close()